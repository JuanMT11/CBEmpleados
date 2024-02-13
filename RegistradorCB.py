import tkinter as tk
from tkinter import messagebox
import cv2
from pyzbar import pyzbar
from openpyxl import Workbook, load_workbook
from datetime import datetime

def decode_barcode(image):
    barcodes = pyzbar.decode(image)
    if len(barcodes) > 0:
        barcode_data = barcodes[0].data.decode("utf-8")
        return barcode_data
    else:
        return None

def registrar_entrada_salida(usuario, registro):
    try:
        workbook = load_workbook("registros_trabajo.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Fecha y Hora", "Usuario", "Acción"])
    
    fecha_hora_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    sheet.append([fecha_hora_actual, usuario, registro])
    workbook.save("registros_trabajo.xlsx")

def scan_and_register():
    usuarios = {}
    try:
        workbook_usuarios = load_workbook("usuarios.xlsx")
        sheet_usuarios = workbook_usuarios.active
        
        for row in sheet_usuarios.iter_rows(min_row=2, values_only=True):
            codigo_barras = row[0]
            nombre = row[1]
            usuarios[codigo_barras] = nombre
    except FileNotFoundError:
        messagebox.showerror("Error", "No se encontró el archivo de usuarios.xlsx")
        return
    
    def handle_scan():
        registro = registro_entry.get()
        if registro.strip() == "":
            messagebox.showerror("Error", "Por favor ingrese la acción (entrada/salida).")
            return
        else:
            camera = cv2.VideoCapture(0)
            while True:
                ret, frame = camera.read()
                
                if not ret:
                    messagebox.showerror("Error", "Error al capturar la imagen.")
                    break
                
                barcode_data = decode_barcode(frame)
                
                if barcode_data is not None:
                    if barcode_data in usuarios:
                        nombre_usuario = usuarios[barcode_data]
                        registrar_entrada_salida(nombre_usuario, registro)
                        messagebox.showinfo("Información", f"Registro de {registro} para {nombre_usuario}")
                        break
                
                cv2.imshow("Barcode Scanner", frame)
                
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
            
            camera.release()
            cv2.destroyAllWindows()
    
    scan_window = tk.Toplevel(root)
    scan_window.title("Escanear y Registrar")
    
    registro_label = tk.Label(scan_window, text="Ingrese la acción (entrada/salida):")
    registro_label.pack()
    
    registro_entry = tk.Entry(scan_window)
    registro_entry.pack()
    
    scan_button = tk.Button(scan_window, text="Escanear", command=handle_scan)
    scan_button.pack()

root = tk.Tk()
root.title("Escáner de Código de Barras")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

scan_button = tk.Button(frame, text="Escanear y Registrar", command=scan_and_register)
scan_button.pack()

root.mainloop()
