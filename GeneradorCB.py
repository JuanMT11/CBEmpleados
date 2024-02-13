import tkinter as tk
from tkinter import messagebox
import barcode
from barcode.writer import ImageWriter
from openpyxl import Workbook, load_workbook

def generar_codigo_barras(datos, tipo, nombre_archivo):
    codigo_barras = barcode.get(tipo, datos, writer=ImageWriter())
    codigo_barras.save(nombre_archivo)

def guardar_en_excel(usuarios):
    try:
        workbook = load_workbook("usuarios.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Código de barras", "Nombre de usuario"])
    
    for codigo, nombre in usuarios.items():
        codigo_existente = False
        for row in sheet.iter_rows(min_row=2, values_only=True, max_col=1):
            if row[0] == codigo:
                codigo_existente = True
                break
        
        if not codigo_existente:
            sheet.append([codigo, nombre])
    
    workbook.save("usuarios.xlsx")

def cargar_registros():
    usuarios = {}
    try:
        workbook = load_workbook("usuarios.xlsx")
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo = row[0]
            nombre = row[1]
            usuarios[codigo] = nombre
    except FileNotFoundError:
        pass
    
    return usuarios

def generar_codigo_y_guardar(usuario, nombre_usuario):
    tipo_codigo = "code128"
    nombre_archivo = f"codigo_barras_{usuario}.png"
    generar_codigo_barras(usuario, tipo_codigo, nombre_archivo)
    usuarios[usuario] = nombre_usuario
    guardar_en_excel({usuario: nombre_usuario})
    messagebox.showinfo("Información", f"Se ha generado un código de barras para {nombre_usuario}")

def agregar_usuario():
    usuario = codigo_entry.get()
    nombre_usuario = nombre_entry.get()
    if usuario.strip() == "" or nombre_usuario.strip() == "":
        messagebox.showerror("Error", "Por favor ingrese el código de barras y el nombre del usuario.")
    else:
        generar_codigo_y_guardar(usuario, nombre_usuario)
        codigo_entry.delete(0, tk.END)
        nombre_entry.delete(0, tk.END)

usuarios = cargar_registros()

root = tk.Tk()
root.title("Generador de Códigos de Barras")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

tk.Label(frame, text="Código de Barras:").grid(row=0, column=0, sticky="w")
codigo_entry = tk.Entry(frame)
codigo_entry.grid(row=0, column=1)

tk.Label(frame, text="Nombre del Usuario:").grid(row=1, column=0, sticky="w")
nombre_entry = tk.Entry(frame)
nombre_entry.grid(row=1, column=1)

agregar_button = tk.Button(frame, text="Agregar Usuario", command=agregar_usuario)
agregar_button.grid(row=2, column=0, columnspan=2, pady=10)

root.mainloop()
